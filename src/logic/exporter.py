"""
Excel Exporter Module

Handles writing formatted tags to Excel files with proper formatting and highlighting.
"""

import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Border, Side, Alignment
from openpyxl.utils.dataframe import dataframe_to_rows
from typing import List
from ..models.tag import Tag


class ExcelExporter:
    """
    Exporter for writing formatted tags to Excel files.
    
    Handles the creation of Excel workbooks with proper formatting,
    highlighting, and layout according to the tag structure specification.
    """
    
    def __init__(self, customization_service=None):
        """
        Initialize the Excel exporter.
        
        Args:
            customization_service: CustomizationService instance for applying custom settings
        """
        self.customization_service = customization_service
        self.workbook = None
        self.worksheet = None
        
        # Define styles
        self.black_border = Border(
            left=Side(style='thin', color='000000'),
            right=Side(style='thin', color='000000'),
            top=Side(style='thin', color='000000'),
            bottom=Side(style='thin', color='000000')
        )
        
        self.white_fill = PatternFill(start_color="FFFFFF", end_color="FFFFFF", fill_type="solid")
        self.light_grey_fill = PatternFill(start_color="D3D3D3", end_color="D3D3D3", fill_type="solid")
        self.very_light_grey_fill = PatternFill(start_color="F5F5F5", end_color="F5F5F5", fill_type="solid")
        self.highlight_fill = PatternFill(start_color="CCCCCC", end_color="CCCCCC", fill_type="solid")
        
        self.bold_font = Font(bold=True)
        self.normal_font = Font(bold=False)
    
    def _get_font(self, bold: bool = False):
        """
        Get a font object based on customization settings.
        
        Args:
            bold: Whether the font should be bold
            
        Returns:
            Font object with custom settings applied
        """
        font_name = "Arial"  # Default
        font_size = 10  # Default
        
        if self.customization_service:
            font_name = self.customization_service.get_font_family()
            font_size = self.customization_service.get_font_size()
        
        return Font(name=font_name, size=font_size, bold=bold)
    
    def _get_bold_font(self):
        """
        Get a bold font object based on customization settings.
        
        Returns:
            Bold Font object with custom settings applied
        """
        return self._get_font(bold=True)
    
    def _get_normal_font(self):
        """
        Get a normal font object based on customization settings.
        
        Returns:
            Normal Font object with custom settings applied
        """
        return self._get_font(bold=False)
    
    def _get_header_fill(self):
        """
        Get the header fill color based on customization settings.
        
        Returns:
            PatternFill object for header background
        """
        if self.customization_service:
            color = self.customization_service.get_header_fill_color()
            # Remove the # prefix if present for Excel
            if color.startswith('#'):
                color = color[1:]
            return PatternFill(start_color=color, end_color=color, fill_type="solid")
        else:
            return self.light_grey_fill
    
    def _get_exceedance_fill(self):
        """
        Get the exceedance fill color based on customization settings.
        
        Returns:
            PatternFill object for exceedance background
        """
        if self.customization_service:
            color = self.customization_service.get_exceedance_fill_color()
            # Remove the # prefix if present for Excel
            if color.startswith('#'):
                color = color[1:]
            return PatternFill(start_color=color, end_color=color, fill_type="solid")
        else:
            return self.highlight_fill
    
    def _get_exceedance_font(self, exceeded_standards_col=None):
        """
        Get the font for exceedance cells based on customization settings.
        
        Args:
            exceeded_standards_col: Name of the standards column that was exceeded (for text color)
            
        Returns:
            Font object with appropriate bold and color settings
        """
        font_name = "Arial"  # Default
        font_size = 10  # Default
        bold = False
        color = None
        
        if self.customization_service:
            font_name = self.customization_service.get_font_family()
            font_size = self.customization_service.get_font_size()
            bold = self.customization_service.get_exceedance_bold()
            
            # Get text color from exceeded standards column if provided
            if exceeded_standards_col:
                text_color = self.customization_service.get_standards_text_color(exceeded_standards_col)
                # Remove the # prefix if present for Excel
                if text_color.startswith('#'):
                    text_color = text_color[1:]
                color = text_color
        
        return Font(name=font_name, size=font_size, bold=bold, color=color)
    
    def _convert_to_appropriate_type(self, value):
        """
        Convert value to appropriate type to avoid Excel green triangles.
        
        Numeric strings are converted to numbers, while non-numeric strings
        (like "< 100 U") remain as strings.
        
        Args:
            value: The value to convert (can be string, int, float, etc.)
            
        Returns:
            Converted value (number if numeric, original value otherwise)
        """
        # If already a number, return as-is
        if isinstance(value, (int, float)):
            return value
        
        # If not a string, return as-is
        if not isinstance(value, str):
            return value
        
        # Empty string stays as string
        if not value.strip():
            return value
        
        # Check if it looks like a non-detect value (starts with "<" or contains letters)
        # These should remain as strings
        value_stripped = value.strip()
        if value_stripped.startswith('<') or value_stripped.startswith('>'):
            return value
        
        # Check if it contains letters (like "ND", "U", etc.) - keep as string
        if any(c.isalpha() for c in value_stripped):
            return value
        
        # Try to convert to number
        try:
            # Try integer first
            if '.' not in value_stripped:
                return int(value_stripped)
            else:
                return float(value_stripped)
        except (ValueError, TypeError):
            # If conversion fails, return as string
            return value
    
    def export_tags(self, tags: List[Tag], output_path: str) -> bool:
        """
        Export a list of tags to an Excel file.
        
        Args:
            tags: List of Tag objects to export
            output_path: Path where the Excel file should be saved
            
        Returns:
            True if export was successful, False otherwise
        """
        try:
            # Create new workbook and worksheet
            self.workbook = Workbook()
            self.worksheet = self.workbook.active
            self.worksheet.title = "Tags"
            
            # Write each tag
            current_row = 1
            for i, tag in enumerate(tags):
                current_row = self._write_tag(tag, current_row)
                
                # Add spacing between tags (except after the last one)
                if i < len(tags) - 1:
                    current_row += 2
            
            # Auto-fit column widths
            self._auto_fit_columns()
            
            # Save the workbook
            self.workbook.save(output_path)
            return True
            
        except Exception as e:
            print(f"Error exporting tags: {e}")
            return False
    
    def _write_tag(self, tag: Tag, start_row: int) -> int:
        """
        Write a single tag to the worksheet.
        
        Args:
            tag: The Tag object to write
            start_row: The starting row number in the worksheet
            
        Returns:
            The next available row number after writing the tag
        """
        current_row = start_row
        all_rows = tag.get_all_rows()
        
        if not all_rows:
            return current_row
        
        # Detect if this is a sample depth format tag (2 columns, first row starts with "Location ID:")
        first_row = all_rows[0]
        is_sample_depth_format = (len(first_row.values) == 2 and 
                                  len(first_row.values) > 0 and 
                                  str(first_row.values[0]).startswith("Location ID:"))
        
        if is_sample_depth_format:
            # Sample depth format: 2-column vertical blocks
            return self._write_tag_sample_depth_format(tag, start_row)
        else:
            # Regular format: column-based with merged header
            return self._write_tag_regular_format(tag, start_row)
    
    def _write_tag_sample_depth_format(self, tag: Tag, start_row: int) -> int:
        """
        Write a tag in sample depth format (2-column vertical blocks).
        
        Args:
            tag: The Tag object to write
            start_row: The starting row number in the worksheet
            
        Returns:
            The next available row number after writing the tag
        """
        current_row = start_row
        all_rows = tag.get_all_rows()
        
        # Write all rows with 2 columns
        for row in all_rows:
            # Write values to the current row (2 columns)
            for col_idx, value in enumerate(row.values[:2], 1):  # Only use first 2 columns
                # Convert date strings to datetime objects for proper Excel formatting
                if col_idx == 2 and str(row.values[0] if len(row.values) > 0 else "") == "Date Sampled:":
                    # This is a date value - try to parse it as datetime
                    from datetime import datetime
                    try:
                        # Try to parse the date string
                        if isinstance(value, str) and len(value) > 10:
                            # Format: "YYYY-MM-DD HH:MM:SS" - extract just the date part
                            date_obj = datetime.strptime(value.split()[0], '%Y-%m-%d')
                            cell = self.worksheet.cell(row=current_row, column=col_idx, value=date_obj)
                            # Set number format to show only date (no time) - Excel format code
                            cell.number_format = 'yyyy-mm-dd'
                        elif isinstance(value, str):
                            # Format: "YYYY-MM-DD"
                            date_obj = datetime.strptime(value, '%Y-%m-%d')
                            cell = self.worksheet.cell(row=current_row, column=col_idx, value=date_obj)
                            # Set number format to show only date (no time) - Excel format code
                            cell.number_format = 'yyyy-mm-dd'
                        else:
                            # Already a datetime object or other type
                            cell = self.worksheet.cell(row=current_row, column=col_idx, value=value)
                            if isinstance(value, datetime):
                                # Set number format to show only date (no time) - Excel format code
                                cell.number_format = 'yyyy-mm-dd'
                    except (ValueError, TypeError):
                        # If parsing fails, write as string (but convert to appropriate type)
                        cell_value = self._convert_to_appropriate_type(value)
                        cell = self.worksheet.cell(row=current_row, column=col_idx, value=cell_value)
                else:
                    # Try to convert numeric strings to numbers to avoid green triangles
                    cell_value = self._convert_to_appropriate_type(value)
                    cell = self.worksheet.cell(row=current_row, column=col_idx, value=cell_value)
                
                # Apply base formatting
                cell.border = self.black_border
                cell.font = self._get_normal_font()
                
                # Determine if this is an analyte row (has analyte_name set)
                is_analyte_row = bool(row.analyte_name)
                
                # Determine if this is an exceedance cell
                is_exceedance = (col_idx == 2 and len(row.is_highlighted) >= col_idx and 
                                row.is_highlighted[col_idx - 1])
                
                # Get exceeded standards column for text color
                exceeded_standards_col = None
                if is_exceedance and hasattr(row, 'exceeded_standards_cols') and row.exceeded_standards_cols:
                    if col_idx - 1 < len(row.exceeded_standards_cols):
                        exceeded_standards_col = row.exceeded_standards_cols[col_idx - 1]
                
                # Apply fill color based on row type and exceedance status
                if is_exceedance:
                    # Exceedance cells get fill color and font from customization settings
                    cell.fill = self._get_exceedance_fill()
                    cell.font = self._get_exceedance_font(exceeded_standards_col)
                elif is_analyte_row:
                    # Analyte rows (analyte name and value) get white fill
                    cell.fill = self.white_fill
                else:
                    # All other cells (labels, location ID, dates, intervals) get very light grey fill
                    cell.fill = self.very_light_grey_fill
                
                # Determine if this is a label row (Location ID, Date Sampled, Sample interval)
                # Left column contains labels ending with ":", right column contains values
                if col_idx == 1:
                    # Left column with labels - left align
                    cell.alignment = Alignment(horizontal='left', vertical='center')
                else:
                    # Right column with values - center align
                    cell.alignment = Alignment(horizontal='center', vertical='center')
            
            current_row += 1
        
        return current_row
    
    def _write_tag_regular_format(self, tag: Tag, start_row: int) -> int:
        """
        Write a tag in regular format (column-based with merged header).
        
        Args:
            tag: The Tag object to write
            start_row: The starting row number in the worksheet
            
        Returns:
            The next available row number after writing the tag
        """
        current_row = start_row
        all_rows = tag.get_all_rows()
        
        # Calculate the total number of columns for this tag
        # This includes: 1 column for analyte names + number of dates
        total_columns = 1 + len(tag.dates)
        
        # Write the first header row (Sample ID) as a merged cell
        if all_rows:
            first_row = all_rows[0]
            if len(first_row.values) > 1:
                # Merge across the entire width of the tag
                self.worksheet.merge_cells(
                    start_row=current_row, start_column=1,
                    end_row=current_row, end_column=total_columns
                )
                
                # Write the sample ID (location_id) in the merged cell
                sample_id_cell = self.worksheet.cell(row=current_row, column=1, value=tag.location_id)
                sample_id_cell.font = self._get_bold_font()
                sample_id_cell.fill = self._get_header_fill()
                sample_id_cell.border = self.black_border
                sample_id_cell.alignment = Alignment(horizontal='center', vertical='center')
                
                # Apply borders to the merged cell properly - ensure top border is visible
                self._apply_merged_cell_borders(current_row, 1, current_row, total_columns)
            
            current_row += 1
        
        # Write the remaining rows (starting from the second row)
        for row_idx, row in enumerate(all_rows[1:], 1):
            # Determine if this is a header row (only the date row is a header)
            # Row 1 = sample ID (merged), Row 2 = date header, Row 3+ = analyte data
            is_header_row = row_idx == 1  # Only the date row is a header
            
            # Write values to the current row
            for col_idx, value in enumerate(row.values, 1):
                # Try to convert numeric strings to numbers to avoid green triangles
                cell_value = self._convert_to_appropriate_type(value)
                cell = self.worksheet.cell(row=current_row, column=col_idx, value=cell_value)
                
                # Apply base formatting
                cell.border = self.black_border
                
                # Apply fill color based on row type
                if is_header_row:
                    cell.fill = self._get_header_fill()
                    cell.font = self._get_bold_font()
                    cell.alignment = Alignment(horizontal='center', vertical='center')
                else:
                    cell.fill = self.white_fill
                    cell.font = self._get_normal_font()
                    
                    # First column (analyte names) - check customization setting for alignment
                    if col_idx == 1:
                        if (self.customization_service and 
                            self.customization_service.get_center_analyte_names()):
                            # Center analyte names if setting is enabled
                            cell.alignment = Alignment(horizontal='center', vertical='center')
                        else:
                            # Default left alignment
                            cell.alignment = Alignment(horizontal='left', vertical='center')
                    else:
                        cell.alignment = Alignment(horizontal='center', vertical='center')
                
                # Apply highlighting for analyte values (only for non-header rows)
                if not is_header_row and col_idx <= len(row.is_highlighted):
                    if row.is_highlighted[col_idx - 1]:
                        # Get exceeded standards column for text color
                        exceeded_standards_col = None
                        if hasattr(row, 'exceeded_standards_cols') and row.exceeded_standards_cols:
                            if col_idx - 1 < len(row.exceeded_standards_cols):
                                exceeded_standards_col = row.exceeded_standards_cols[col_idx - 1]
                        
                        # Apply exceedance formatting
                        cell.font = self._get_exceedance_font(exceeded_standards_col)
                        cell.fill = self._get_exceedance_fill()
            
            current_row += 1
        
        return current_row
    
    def _apply_merged_cell_borders(self, start_row, start_col, end_row, end_col):
        """
        Apply borders to a merged cell by setting borders on the corner cells.
        
        Args:
            start_row: Starting row of the merged cell
            start_col: Starting column of the merged cell
            end_row: Ending row of the merged cell
            end_col: Ending column of the merged cell
        """
        # Apply borders to the corner cells of the merged range
        # Top-left corner
        top_left = self.worksheet.cell(row=start_row, column=start_col)
        top_left.border = Border(
            left=Side(style='thin', color='000000'),
            top=Side(style='thin', color='000000')
        )
        
        # Top-right corner
        top_right = self.worksheet.cell(row=start_row, column=end_col)
        top_right.border = Border(
            right=Side(style='thin', color='000000'),
            top=Side(style='thin', color='000000')
        )
        
        # Bottom-left corner
        bottom_left = self.worksheet.cell(row=end_row, column=start_col)
        bottom_left.border = Border(
            left=Side(style='thin', color='000000'),
            bottom=Side(style='thin', color='000000')
        )
        
        # Bottom-right corner
        bottom_right = self.worksheet.cell(row=end_row, column=end_col)
        bottom_right.border = Border(
            right=Side(style='thin', color='000000'),
            bottom=Side(style='thin', color='000000')
        )
        
        # Also apply borders to the merged cell itself to ensure top border is visible
        merged_cell = self.worksheet.cell(row=start_row, column=start_col)
        merged_cell.border = self.black_border
        
        # Apply borders to all cells in the merged range to ensure complete border
        for col in range(start_col, end_col + 1):
            cell = self.worksheet.cell(row=start_row, column=col)
            cell.border = self.black_border
    
    def _auto_fit_columns(self):
        """
        Auto-fit column widths based on content.
        """
        from openpyxl.utils import get_column_letter
        
        # Iterate through all columns
        for col_idx, column in enumerate(self.worksheet.columns, 1):
            max_length = 0
            column_letter = get_column_letter(col_idx)
            
            # Find the maximum content length in this column
            for cell in column:
                try:
                    if cell.value is not None:
                        # Convert cell value to string and get length
                        cell_value = str(cell.value)
                        # For dates, use the formatted display length
                        if hasattr(cell, 'number_format') and cell.number_format:
                            # If it's a date format, estimate display length
                            if 'yyyy' in cell.number_format.lower() or 'mm' in cell.number_format.lower():
                                cell_value = 'YYYY-MM-DD'  # Approximate date display length
                        cell_length = len(cell_value)
                        if cell_length > max_length:
                            max_length = cell_length
                except (TypeError, AttributeError):
                    pass
            
            # Set column width with padding (minimum 10, maximum 50)
            # Add extra padding for better readability
            adjusted_width = min(max(max_length + 2, 10), 50)
            self.worksheet.column_dimensions[column_letter].width = adjusted_width
    
    def _apply_cell_formatting(self, cell, is_highlighted: bool):
        """
        Apply formatting to a cell based on highlighting status.
        
        Args:
            cell: The worksheet cell to format
            is_highlighted: Whether the cell should be highlighted
        """
        if is_highlighted:
            # Apply bold font and grey background for highlighted cells
            cell.font = Font(bold=True)
            cell.fill = PatternFill(start_color="CCCCCC", end_color="CCCCCC", fill_type="solid")
        # Leave unhighlighted cells with default formatting
    
    def _create_header_style(self):
        """
        Create the style for header rows.
        
        Returns:
            Font style for headers
        """
        # TODO: Implement header styling
        # - Return bold font style for header rows
        pass 